import os
import qrcode
from openpyxl import load_workbook
from PIL import ImageDraw, ImageFont

# get current directory
cwd = os.getcwd()

# load the Arial font with a specific size
font = ImageFont.truetype("arial.ttf", 20)

# loop through all files in current directory
for file in os.listdir(cwd):
    # check if file is an excel file
    if file.endswith(".xlsx"):
        # load excel workbook
        wb = load_workbook(file)
        # loop through all worksheets in workbook
        for ws in wb.worksheets:
            # loop through all rows in worksheet
            for row in ws.iter_rows(min_row=2, values_only=True):
                # get barcode value from row and convert to string
                barcode = str(row[2])
                # generate QR code
                qr = qrcode.QRCode(version=1, box_size=10, border=5)
                qr.add_data(barcode)
                qr.make(fit=True)
                img = qr.make_image(fill_color="black", back_color="white")
                # add barcode number below QR code
                img_with_text = img.copy()
                text_draw = ImageDraw.Draw(img_with_text)
                
                # Calculate the position for the barcode text
                text_bbox = font.getbbox(barcode)
                text_width, text_height = text_bbox[2] - text_bbox[0], text_bbox[3] - text_bbox[1]
                img_width, img_height = img_with_text.size
                text_x = (img_width - text_width) // 2
                text_y = img_height - text_height
                
                text_draw.text((text_x, text_y), barcode, fill='black', font=font)
                
                # save QR code image as png or jpg
                filename = f"{barcode.encode('utf-8').hex()}.png"  # Use hex representation for the filename
                img_with_text.save(filename)
